import csv
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook
import openpyxl
import datetime

DEBUG = False

def printTransactions(array, info):
    print(info)
    for transaction in array:
        print(transaction.__dict__)


def getFileNameByDialog():
    Tk().withdraw()
    fileName = askopenfilename()
    return fileName


def openPortfolioWorkbookByDialog():
    fileName = getFileNameByDialog()
    return load_workbook(filename=fileName)


def openTransactionWorkbookByDialog():
    fileName = getFileNameByDialog()
    file = open(fileName)
    return csv.reader(file)


def writeWithNewLine(writer, message):
    writer.write(message)
    writer.write('\n')


def writeToFile(writer, message, sheet=None, column=None, row=None, value=None, dateFormat=None, moneyFormat=None,
                percentFormat=None, color=None, twoDecimalsFormat=None):
    if DEBUG == True:
        writeToLog(writer, message, sheet, column, row, value)
    if sheet and column and row and value:
        sheet.cell(column=column, row=row, value=value)
        if dateFormat:
            sheet.cell(
                row=row, column=column).number_format = 'mm/dd/yyyy'
        if moneyFormat:
            sheet.cell(
                row=row, column=column).number_format = '"$"#,##0.00_-'
        if percentFormat:
            sheet.cell(
                row=row, column=column).number_format = '0.00%'
        if color:
            yellow = openpyxl.styles.colors.Color(rgb=color)
            filling = openpyxl.styles.fills.PatternFill(
                patternType='solid', fgColor=yellow)
            sheet.cell(
                row=row, column=column).fill = filling
        if twoDecimalsFormat:
            sheet.cell(
                row=row, column=column).number_format = '0.00'


def writeToLog(writer, message, sheet, column, row, value):
    writeWithNewLine(writer, message)
    if sheet and column and row and value:
        writeWithNewLine(writer, 'sheet: ' + str(sheet))
        writeWithNewLine(writer, 'column: ' + str(column))
        writeWithNewLine(writer, 'row: ' + str(row))
        writeWithNewLine(writer, 'value: ' + str(value))
    writeWithNewLine(writer, '--------------------')



def getMonthNumber(monthSpelledOut):
    datetime_object = datetime.datetime.strptime(monthSpelledOut, "%b")
    return datetime_object.month


def getValueFormula(symbol):
    return '=RTD("tos.rtd",,"last", "' + symbol + '")'


def getAskFormula(newTransaction):
    base = '=RTD("tos.rtd",,"ASK",".'
    strike = str(newTransaction.strike)
    strike = strike.replace('.0', '')
    day = str(newTransaction.day)
    if len(day) == 1:
        day = '0' + day
    month = str(getMonthNumber(newTransaction.month))
    if len(month) == 1:
        month = '0' + month
    year = str(newTransaction.year)[2:]
    symbolDate = newTransaction.symbol + year + \
        month + day + 'P' + strike
    return base + symbolDate + '")'


def getCallFormula(newTransaction):
    base = '=RTD("tos.rtd",,"LAST",".'
    strike = str(newTransaction.strike)
    strike = strike.replace('.0', '')
    day = str(newTransaction.day)
    if len(day) == 1:
        day = '0' + day
    month = str(getMonthNumber(newTransaction.month))
    if len(month) == 1:
        month = '0' + month
    year = str(newTransaction.year)[2:]
    symbolDate = newTransaction.symbol + year + \
        month + day + 'C' + strike
    return base + symbolDate + '")'


def printWithNewLines(message):
    print('\n' + message + '\n')


def getWriteableTransactions(transactions):
    writeableNewTransactions = []
    for transaction in transactions:
        if transaction.type != "REMOVAL":
            writeableNewTransactions.append(transaction)
    return writeableNewTransactions
